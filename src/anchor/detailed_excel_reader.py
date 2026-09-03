"""Detailed Operating Model V2.1 Gate 10 -- Excel ingestion for Detailed
Underwrite.

Parallels ``excel_reader.py``'s structure and rigor exactly (same ``Inputs``
worksheet/table shape, same deterministic error contract via
``InputIssue``/``InputValidationError``, same reuse of the shared
``_normalize_field_value``-family domain rules and the authoritative
``validate_acquisition_terms``/``validate_detailed_operating_inputs``
constructors) over the Detailed workbook's own field set: the eleven
``AcquisitionTerms`` fields plus the eleven ``DetailedOperatingInputs``
fields, twenty-two Field IDs total, every one of them required -- there is
no legacy Detailed workbook and therefore no optional/defaulted field here.

This module never calculates NOI, acquisition results, or any other
financial output -- it produces proposed assumptions only. It is a distinct
reader/contract from Quick's, not a variant of it: a Detailed workbook is
identified by its own explicit ``anchor_schema``/``schema_version`` metadata
(``anchor.workbook_schema``), never by which Field IDs happen to be present.
"""

from __future__ import annotations

from dataclasses import dataclass, replace
from io import BytesIO
from numbers import Real
from os import PathLike
from typing import Any

from openpyxl import load_workbook

from .contracts import AcquisitionTerms, DetailedOperatingInputs
from .excel_reader import (
    _last_content_row,
    _malformed_field_id_issue,
    _validate_header,
    _workbook_identifier,
    _workbook_open_error,
)
from .validation import (
    DETAILED_FIELD_IDS,
    TERMS_FIELD_IDS,
    InputIssue,
    InputValidationError,
    IssueCategory,
    _normalize_detailed_field_value,
    _normalize_field_value,
    validate_acquisition_terms,
    validate_detailed_operating_inputs,
)
from .workbook_schema import (
    DETAILED_SCHEMA,
    QUICK_SCHEMA,
    SUPPORTED_DETAILED_SCHEMA_VERSION,
    _is_blank,
    read_workbook_schema,
)

_INPUTS_SHEET = "Inputs"
_TEXT_CELL_TYPES = frozenset({"s", "inlineStr"})

#: Every Detailed Field ID, in canonical display order (AcquisitionTerms
#: first, then DetailedOperatingInputs) -- drives duplicate/missing
#: detection and value-issue ordering, mirroring ``ALL_FIELD_IDS``'s role
#: in the Quick reader.
_DETAILED_WORKBOOK_FIELD_IDS = TERMS_FIELD_IDS + DETAILED_FIELD_IDS
_TERMS_FIELD_ID_SET = frozenset(TERMS_FIELD_IDS)


@dataclass(frozen=True, slots=True, kw_only=True)
class DetailedExcelIntakeReport:
    """The Detailed counterpart to ``ExcelIntakeReport``: proposed
    assumptions only, plus the workbook's own declared schema/version --
    never an ``OperatingProjection``, ``AcquisitionResults``, or any other
    calculated value. Every field on both contracts is always present --
    unlike Quick's ``defaulted_v2_field_ids``, there is no optional Detailed
    field to report as defaulted."""

    terms: AcquisitionTerms
    detailed_operating_inputs: DetailedOperatingInputs
    anchor_schema: str
    schema_version: str


def read_detailed_excel_intake(
    workbook_path: str | PathLike[str],
) -> DetailedExcelIntakeReport:
    """Read one canonical Detailed ``.xlsx`` workbook file into validated
    ``AcquisitionTerms`` + ``DetailedOperatingInputs``, plus its declared
    schema/version."""

    identifier = _workbook_identifier(workbook_path)
    if not identifier.casefold().endswith(".xlsx"):
        raise _workbook_open_error(identifier)
    return _read_detailed_excel_intake_from_source(workbook_path, identifier)


def read_detailed_excel_intake_from_bytes(data: bytes) -> DetailedExcelIntakeReport:
    """Like :func:`read_detailed_excel_intake`, but for an in-memory
    workbook (e.g. an HTTP upload)."""

    return _read_detailed_excel_intake_from_source(BytesIO(data), "<uploaded workbook>")


def _schema_mismatch(message: str) -> InputValidationError:
    return InputValidationError(
        (InputIssue(category=IssueCategory.SCHEMA_MISMATCH, message=message),)
    )


def _read_detailed_excel_intake_from_source(
    source: str | PathLike[str] | BytesIO,
    identifier: str,
) -> DetailedExcelIntakeReport:
    try:
        workbook = load_workbook(source, data_only=False)
    except Exception:
        raise _workbook_open_error(identifier) from None

    try:
        # Gate 10 wrong-workbook protection: a Detailed workbook always
        # declares itself explicitly. Absent metadata, or an explicit Quick
        # declaration, means this is a Quick workbook (legacy schema-less
        # Quick workbooks carry no ``Meta`` sheet at all) -- never flattened
        # into Detailed assumptions by coincidental field overlap (both
        # modes share eleven ``AcquisitionTerms``-shaped Field IDs).
        schema = read_workbook_schema(workbook)
        if schema.anchor_schema is None or schema.anchor_schema == QUICK_SCHEMA:
            raise _schema_mismatch(
                "This workbook uses the Quick Underwrite schema. Switch to "
                "Quick Underwrite or upload a Detailed Underwrite workbook."
            )
        if schema.anchor_schema != DETAILED_SCHEMA:
            raise _schema_mismatch(
                f"Unsupported workbook schema {schema.anchor_schema!r} for "
                "Detailed Underwrite ingestion."
            )
        if schema.schema_version != SUPPORTED_DETAILED_SCHEMA_VERSION:
            raise InputValidationError(
                (
                    InputIssue(
                        category=IssueCategory.UNSUPPORTED_SCHEMA_VERSION,
                        message=(
                            f"Unsupported Detailed workbook schema_version "
                            f"{schema.schema_version!r}. This version of Anchor "
                            f"supports schema_version "
                            f"{SUPPORTED_DETAILED_SCHEMA_VERSION!r}."
                        ),
                    ),
                )
            )

        if _INPUTS_SHEET not in workbook.sheetnames:
            raise InputValidationError(
                (
                    InputIssue(
                        category=IssueCategory.MISSING_SHEET,
                        message=(
                            "Workbook is missing the required exactly named "
                            f"'{_INPUTS_SHEET}' worksheet."
                        ),
                    ),
                )
            )

        worksheet = workbook[_INPUTS_SHEET]
        header_issue = _validate_header(worksheet)
        if header_issue is not None:
            raise InputValidationError((header_issue,))

        row_issues: list[InputIssue] = []
        records: dict[str, list[tuple[int, Any]]] = {
            field_id: [] for field_id in _DETAILED_WORKBOOK_FIELD_IDS
        }

        for row_number in range(2, _last_content_row(worksheet) + 1):
            cells = tuple(worksheet.cell(row=row_number, column=column) for column in range(1, 5))
            if all(_is_blank(cell.value) for cell in cells):
                continue

            field_cell = cells[0]
            field_id = field_cell.value
            if field_cell.data_type == "f":
                row_issues.append(
                    _malformed_field_id_issue(
                        row_number,
                        "Field ID must be literal text and cannot be a formula",
                    )
                )
            elif _is_blank(field_id):
                row_issues.append(
                    _malformed_field_id_issue(row_number, "Field ID is blank")
                )
            elif (
                not isinstance(field_id, str)
                or field_cell.data_type not in _TEXT_CELL_TYPES
            ):
                row_issues.append(
                    _malformed_field_id_issue(
                        row_number, "Field ID must be literal text"
                    )
                )
            elif field_id not in records:
                row_issues.append(
                    InputIssue(
                        category=IssueCategory.UNKNOWN_FIELD_ID,
                        message=(
                            f"Unknown Field ID {field_id!r} at "
                            f"{_INPUTS_SHEET}!A{row_number}."
                        ),
                        field_id=field_id,
                        row=row_number,
                        cell=f"A{row_number}",
                        value=field_id,
                    )
                )
            else:
                records[field_id].append((row_number, cells[2]))

        duplicate_issues = _duplicate_issues(records)
        missing_issues = _missing_issues(records)
        value_issues, normalized = _normalize_unique_values(records)

        issues = tuple(row_issues + duplicate_issues + missing_issues + value_issues)
        if issues:
            raise InputValidationError(issues)

        terms_values = {field_id: normalized[field_id] for field_id in TERMS_FIELD_IDS}
        detailed_values = {
            field_id: normalized[field_id] for field_id in DETAILED_FIELD_IDS
        }
        terms = validate_acquisition_terms(terms_values)
        detailed_operating_inputs = validate_detailed_operating_inputs(detailed_values)

        assert schema.anchor_schema is not None
        assert schema.schema_version is not None
        return DetailedExcelIntakeReport(
            terms=terms,
            detailed_operating_inputs=detailed_operating_inputs,
            anchor_schema=schema.anchor_schema,
            schema_version=schema.schema_version,
        )
    except InputValidationError:
        raise
    except Exception:
        raise _workbook_open_error(identifier) from None
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _duplicate_issues(records: dict[str, list[tuple[int, Any]]]) -> list[InputIssue]:
    issues: list[InputIssue] = []
    for field_id in _DETAILED_WORKBOOK_FIELD_IDS:
        occurrences = records[field_id]
        if len(occurrences) <= 1:
            continue
        rows = tuple(row_number for row_number, _ in occurrences)
        issues.append(
            InputIssue(
                category=IssueCategory.DUPLICATE_FIELD_ID,
                message=f"Duplicate Field ID '{field_id}' occurs on rows {rows}.",
                field_id=field_id,
                rows=rows,
            )
        )
    return issues


def _missing_issues(records: dict[str, list[tuple[int, Any]]]) -> list[InputIssue]:
    return [
        InputIssue(
            category=IssueCategory.MISSING_FIELD_ID,
            message=f"Missing required Field ID '{field_id}'.",
            field_id=field_id,
        )
        for field_id in _DETAILED_WORKBOOK_FIELD_IDS
        if not records[field_id]
    ]


def _normalize_unique_values(
    records: dict[str, list[tuple[int, Any]]],
) -> tuple[list[InputIssue], dict[str, float | int]]:
    issues: list[InputIssue] = []
    normalized: dict[str, float | int] = {}

    for field_id in _DETAILED_WORKBOOK_FIELD_IDS:
        occurrences = records[field_id]
        if len(occurrences) != 1:
            continue

        row_number, value_cell = occurrences[0]
        coordinate = value_cell.coordinate
        raw_value = value_cell.value

        if value_cell.data_type == "f":
            issues.append(
                InputIssue(
                    category=IssueCategory.FORMULA_VALUE,
                    message=(
                        f"Value for Field ID '{field_id}' at "
                        f"{_INPUTS_SHEET}!{coordinate} must be literal and cannot "
                        "be a formula."
                    ),
                    field_id=field_id,
                    row=row_number,
                    cell=coordinate,
                    value=raw_value,
                )
            )
            continue

        if _is_blank(raw_value):
            issues.append(
                InputIssue(
                    category=IssueCategory.BLANK_VALUE,
                    message=(
                        f"Value for Field ID '{field_id}' is blank at "
                        f"{_INPUTS_SHEET}!{coordinate}."
                    ),
                    field_id=field_id,
                    row=row_number,
                    cell=coordinate,
                    value=raw_value,
                )
            )
            continue

        if (
            value_cell.data_type != "n"
            or value_cell.is_date
            or isinstance(raw_value, bool)
            or not isinstance(raw_value, Real)
        ):
            issues.append(
                InputIssue(
                    category=IssueCategory.NON_NUMERIC_VALUE,
                    message=(
                        f"Value for Field ID '{field_id}' at "
                        f"{_INPUTS_SHEET}!{coordinate} must be a literal Excel "
                        "numeric value."
                    ),
                    field_id=field_id,
                    row=row_number,
                    cell=coordinate,
                    value=raw_value,
                )
            )
            continue

        normalize = (
            _normalize_field_value
            if field_id in _TERMS_FIELD_ID_SET
            else _normalize_detailed_field_value
        )
        normalized_value, issue = normalize(field_id, raw_value)
        if issue is not None:
            issues.append(
                replace(
                    issue,
                    message=(
                        f"{issue.message.rstrip('.')} at "
                        f"{_INPUTS_SHEET}!{coordinate}."
                    ),
                    row=row_number,
                    cell=coordinate,
                )
            )
            continue

        assert normalized_value is not None
        normalized[field_id] = normalized_value

    return issues, normalized
