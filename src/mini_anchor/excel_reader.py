"""Canonical Excel ingestion for Mini-Anchor acquisition inputs."""

from __future__ import annotations

from dataclasses import replace
from numbers import Real
from os import PathLike, fsdecode, fspath
from typing import Any

from openpyxl import load_workbook

from .contracts import AcquisitionInputs
from .validation import (
    FIELD_IDS,
    InputIssue,
    InputValidationError,
    IssueCategory,
    _normalize_field_value,
    validate_acquisition_inputs,
)

_INPUTS_SHEET = "Inputs"
_HEADERS = ("Field ID", "Input", "Value", "Unit")
_TEXT_CELL_TYPES = frozenset({"s", "inlineStr"})


def read_acquisition_inputs(
    workbook_path: str | PathLike[str],
) -> AcquisitionInputs:
    """Read one canonical ``.xlsx`` workbook into validated acquisition inputs.

    All workbook and ingestion failures are translated into an ordered
    :class:`InputValidationError` collection. Formulas are intentionally loaded
    as formulas so a cached result can never be mistaken for a literal value.
    """

    identifier = _workbook_identifier(workbook_path)
    if not identifier.casefold().endswith(".xlsx"):
        raise _workbook_open_error(identifier)
    try:
        workbook = load_workbook(workbook_path, data_only=False)
    except Exception:
        raise _workbook_open_error(identifier) from None

    try:
        if _INPUTS_SHEET not in workbook.sheetnames:
            raise InputValidationError(
                (
                    InputIssue(
                        category=IssueCategory.MISSING_SHEET,
                        message=(
                            "Workbook is missing the required exactly named "
                            f"'{_INPUTS_SHEET}' worksheet."
                        ),
                    ),
                )
            )

        worksheet = workbook[_INPUTS_SHEET]
        header_issue = _validate_header(worksheet)
        if header_issue is not None:
            raise InputValidationError((header_issue,))

        row_issues: list[InputIssue] = []
        records: dict[str, list[tuple[int, Any]]] = {
            field_id: [] for field_id in FIELD_IDS
        }

        for row_number in range(2, _last_content_row(worksheet) + 1):
            cells = tuple(worksheet.cell(row=row_number, column=column) for column in range(1, 5))
            if all(_is_blank(cell.value) for cell in cells):
                continue

            field_cell = cells[0]
            field_id = field_cell.value
            if field_cell.data_type == "f":
                row_issues.append(
                    _malformed_field_id_issue(
                        row_number,
                        "Field ID must be literal text and cannot be a formula",
                    )
                )
            elif _is_blank(field_id):
                row_issues.append(
                    _malformed_field_id_issue(row_number, "Field ID is blank")
                )
            elif (
                not isinstance(field_id, str)
                or field_cell.data_type not in _TEXT_CELL_TYPES
            ):
                row_issues.append(
                    _malformed_field_id_issue(
                        row_number, "Field ID must be literal text"
                    )
                )
            elif field_id not in records:
                row_issues.append(
                    InputIssue(
                        category=IssueCategory.UNKNOWN_FIELD_ID,
                        message=(
                            f"Unknown Field ID {field_id!r} at "
                            f"{_INPUTS_SHEET}!A{row_number}."
                        ),
                        field_id=field_id,
                        row=row_number,
                        cell=f"A{row_number}",
                        value=field_id,
                    )
                )
            else:
                records[field_id].append((row_number, cells[2]))

        duplicate_issues = _duplicate_issues(records)
        missing_issues = _missing_issues(records)
        value_issues, normalized = _normalize_unique_values(records)

        issues = tuple(row_issues + duplicate_issues + missing_issues + value_issues)
        if issues:
            raise InputValidationError(issues)

        return validate_acquisition_inputs(normalized)
    except InputValidationError:
        raise
    except Exception:
        raise _workbook_open_error(identifier) from None
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _workbook_open_error(identifier: str) -> InputValidationError:
    return InputValidationError(
        (
            InputIssue(
                category=IssueCategory.WORKBOOK_OPEN,
                message=f"Workbook could not be opened: {identifier}.",
            ),
        )
    )


def _workbook_identifier(workbook_path: object) -> str:
    try:
        identifier = fspath(workbook_path)
        return fsdecode(identifier)
    except Exception:
        try:
            return repr(workbook_path)
        except Exception:
            return "<workbook>"


def _validate_header(worksheet: Any) -> InputIssue | None:
    merged_ranges = sorted(
        worksheet.merged_cells.ranges,
        key=lambda merged_range: (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        ),
    )
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= 1 <= merged_range.max_row
            and merged_range.min_col <= 4
            and merged_range.max_col >= 1
        ):
            return InputIssue(
                category=IssueCategory.MALFORMED_TABLE,
                message=(
                    f"Malformed {_INPUTS_SHEET} header: merged range "
                    f"{merged_range.coord} intersects A1:D1; expected exactly "
                    f"{_HEADERS!r}."
                ),
                row=1,
                cell=merged_range.coord,
            )

    actual = tuple(worksheet.cell(row=1, column=column).value for column in range(1, 5))
    if actual == _HEADERS:
        return None

    first_mismatch = next(
        column
        for column, (actual_value, expected_value) in enumerate(
            zip(actual, _HEADERS, strict=True), start=1
        )
        if actual_value != expected_value
    )
    coordinate = worksheet.cell(row=1, column=first_mismatch).coordinate
    return InputIssue(
        category=IssueCategory.MALFORMED_TABLE,
        message=(
            f"Malformed {_INPUTS_SHEET} header at {coordinate}: expected "
            f"A1:D1 to equal {_HEADERS!r}, found {actual!r}."
        ),
        row=1,
        cell=coordinate,
    )


def _last_content_row(worksheet: Any) -> int:
    return max(
        (
            cell.row
            for cell in worksheet._cells.values()
            if cell.row >= 2 and 1 <= cell.column <= 4 and cell.value is not None
        ),
        default=1,
    )


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _malformed_field_id_issue(row_number: int, detail: str) -> InputIssue:
    return InputIssue(
        category=IssueCategory.MALFORMED_TABLE,
        message=f"Malformed row {row_number}: {detail} at {_INPUTS_SHEET}!A{row_number}.",
        row=row_number,
        cell=f"A{row_number}",
    )


def _duplicate_issues(records: dict[str, list[tuple[int, Any]]]) -> list[InputIssue]:
    issues: list[InputIssue] = []
    for field_id in FIELD_IDS:
        occurrences = records[field_id]
        if len(occurrences) <= 1:
            continue
        rows = tuple(row_number for row_number, _ in occurrences)
        issues.append(
            InputIssue(
                category=IssueCategory.DUPLICATE_FIELD_ID,
                message=f"Duplicate Field ID '{field_id}' occurs on rows {rows}.",
                field_id=field_id,
                rows=rows,
            )
        )
    return issues


def _missing_issues(records: dict[str, list[tuple[int, Any]]]) -> list[InputIssue]:
    return [
        InputIssue(
            category=IssueCategory.MISSING_FIELD_ID,
            message=f"Missing required Field ID '{field_id}'.",
            field_id=field_id,
        )
        for field_id in FIELD_IDS
        if not records[field_id]
    ]


def _normalize_unique_values(
    records: dict[str, list[tuple[int, Any]]],
) -> tuple[list[InputIssue], dict[str, float | int]]:
    issues: list[InputIssue] = []
    normalized: dict[str, float | int] = {}

    for field_id in FIELD_IDS:
        occurrences = records[field_id]
        if len(occurrences) != 1:
            continue

        row_number, value_cell = occurrences[0]
        coordinate = value_cell.coordinate
        raw_value = value_cell.value

        if value_cell.data_type == "f":
            issues.append(
                InputIssue(
                    category=IssueCategory.FORMULA_VALUE,
                    message=(
                        f"Value for Field ID '{field_id}' at "
                        f"{_INPUTS_SHEET}!{coordinate} must be literal and cannot "
                        "be a formula."
                    ),
                    field_id=field_id,
                    row=row_number,
                    cell=coordinate,
                    value=raw_value,
                )
            )
            continue

        if _is_blank(raw_value):
            issues.append(
                InputIssue(
                    category=IssueCategory.BLANK_VALUE,
                    message=(
                        f"Value for Field ID '{field_id}' is blank at "
                        f"{_INPUTS_SHEET}!{coordinate}."
                    ),
                    field_id=field_id,
                    row=row_number,
                    cell=coordinate,
                    value=raw_value,
                )
            )
            continue

        if (
            value_cell.data_type != "n"
            or value_cell.is_date
            or isinstance(raw_value, bool)
            or not isinstance(raw_value, Real)
        ):
            issues.append(
                InputIssue(
                    category=IssueCategory.NON_NUMERIC_VALUE,
                    message=(
                        f"Value for Field ID '{field_id}' at "
                        f"{_INPUTS_SHEET}!{coordinate} must be a literal Excel "
                        "numeric value."
                    ),
                    field_id=field_id,
                    row=row_number,
                    cell=coordinate,
                    value=raw_value,
                )
            )
            continue

        normalized_value, issue = _normalize_field_value(field_id, raw_value)
        if issue is not None:
            issues.append(
                replace(
                    issue,
                    message=(
                        f"{issue.message.rstrip('.')} at "
                        f"{_INPUTS_SHEET}!{coordinate}."
                    ),
                    row=row_number,
                    cell=coordinate,
                )
            )
            continue

        assert normalized_value is not None
        normalized[field_id] = normalized_value

    return issues, normalized
