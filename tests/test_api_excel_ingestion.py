"""Tests for the Phase 10B Excel ingestion endpoint (``POST /ingestion/excel``)
in ``anchor.api``.

Mirrors ``test_api_ingestion.py``'s style. Unlike OM ingestion, there is no
external provider to mock: ``read_acquisition_inputs_from_bytes_with_report`` is the
same deterministic, frozen Phase 1 parsing/validation path the CLI has
always used (shared with ``read_acquisition_inputs`` -- see
``tests/test_excel_reader.py`` for the parity tests proving that), so most
tests here exercise the real function rather than a mock, and only the
delegation test below stubs it out to prove the route doesn't reimplement
any parsing itself.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from anchor.api import app
from anchor.contracts import AcquisitionInputs
from anchor.validation import InputValidationError, IssueCategory

EXAMPLE_WORKBOOK = Path(__file__).resolve().parents[1] / "examples" / "anchor_input.xlsx"

HEADERS = ("Field ID", "Input", "Value", "Unit")
VALUES: dict[str, int | float] = {
    "purchase_price": 50_000_000,
    "current_noi": 2_500_000,
    "occupancy": 0.95,
    "noi_growth": 0.03,
    "hold_period": 5,
    "exit_cap_rate": 0.055,
    "ltv": 0.65,
    "interest_rate": 0.0525,
    "amortization": 30,
}
LABELS = {
    "purchase_price": "Purchase Price",
    "current_noi": "Current NOI",
    "occupancy": "Occupancy",
    "noi_growth": "NOI Growth",
    "hold_period": "Hold Period",
    "exit_cap_rate": "Exit Cap Rate",
    "ltv": "LTV",
    "interest_rate": "Interest Rate",
    "amortization": "Amortization",
}
UNITS = {
    "purchase_price": "USD",
    "current_noi": "USD/year",
    "occupancy": "%",
    "noi_growth": "%/year",
    "hold_period": "years",
    "exit_cap_rate": "%",
    "ltv": "%",
    "interest_rate": "%",
    "amortization": "years",
}
FIELD_ORDER = tuple(VALUES.keys())

EXPECTED = AcquisitionInputs(
    purchase_price=50_000_000.0,
    current_noi=2_500_000.0,
    occupancy=0.95,
    noi_growth=0.03,
    hold_period=5,
    exit_cap_rate=0.055,
    ltv=0.65,
    interest_rate=0.0525,
    amortization=30,
)


@pytest.fixture()
def client() -> TestClient:
    return TestClient(app)


def _build_workbook_bytes(*, values: dict[str, object] | None = None) -> bytes:
    supplied = VALUES if values is None else values
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet.append(HEADERS)
    for field_id in FIELD_ORDER:
        worksheet.append((field_id, LABELS[field_id], supplied[field_id], UNITS[field_id]))
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


VALID_WORKBOOK_BYTES = _build_workbook_bytes()


def _upload_files(content: bytes, filename: str = "anchor_input.xlsx") -> dict[str, Any]:
    return {
        "file": (
            filename,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


# =============================================================================
# Happy path
# =============================================================================


# Underwriting V2 Gate 5: absent from a legacy nine-field workbook, so each
# both defaults to its neutral value on ``inputs`` and is named in the
# response's ``defaulted_v2_field_ids``.
V2_NEUTRAL_DEFAULTS: dict[str, Any] = {
    "acquisition_cost_pct": 0.0,
    "financing_fee_pct": 0.0,
    "disposition_cost_pct": 0.0,
    "annual_capex_reserve": 0.0,
    "io_period": 0,
}
V2_FIELD_IDS = tuple(V2_NEUTRAL_DEFAULTS)


def test_valid_upload_returns_200_with_the_nine_validated_inputs(client: TestClient) -> None:
    response = client.post("/ingestion/excel", files=_upload_files(VALID_WORKBOOK_BYTES))

    assert response.status_code == 200
    body = response.json()
    assert body == {
        "inputs": {
            "purchase_price": 50_000_000.0,
            "current_noi": 2_500_000.0,
            "occupancy": 0.95,
            "noi_growth": 0.03,
            "hold_period": 5,
            "exit_cap_rate": 0.055,
            "ltv": 0.65,
            "interest_rate": 0.0525,
            "amortization": 30,
            **V2_NEUTRAL_DEFAULTS,
        },
        "defaulted_v2_field_ids": list(V2_FIELD_IDS),
    }


def test_valid_upload_of_the_tracked_example_workbook_matches_cli_golden_values(
    client: TestClient,
) -> None:
    """Directly proves the acceptance criterion that the same values the
    existing CLI Excel ingestion produces appear correctly through the web
    upload path -- uses the same tracked fixture and expected values as
    ``tests/test_excel_reader.py``."""

    response = client.post(
        "/ingestion/excel",
        files=_upload_files(EXAMPLE_WORKBOOK.read_bytes()),
    )

    assert response.status_code == 200
    assert response.json() == {
        "inputs": {
            "purchase_price": EXPECTED.purchase_price,
            "current_noi": EXPECTED.current_noi,
            "occupancy": EXPECTED.occupancy,
            "noi_growth": EXPECTED.noi_growth,
            "hold_period": EXPECTED.hold_period,
            "exit_cap_rate": EXPECTED.exit_cap_rate,
            "ltv": EXPECTED.ltv,
            "interest_rate": EXPECTED.interest_rate,
            "amortization": EXPECTED.amortization,
            "acquisition_cost_pct": EXPECTED.acquisition_cost_pct,
            "financing_fee_pct": EXPECTED.financing_fee_pct,
            "disposition_cost_pct": EXPECTED.disposition_cost_pct,
            "annual_capex_reserve": EXPECTED.annual_capex_reserve,
            "io_period": EXPECTED.io_period,
        },
        "defaulted_v2_field_ids": list(V2_FIELD_IDS),
    }


def test_endpoint_delegates_to_the_shared_bytes_reader_rather_than_reimplementing_parsing(
    client: TestClient,
) -> None:
    from anchor.excel_reader import ExcelIntakeReport

    stub_report = ExcelIntakeReport(inputs=EXPECTED, defaulted_v2_field_ids=V2_FIELD_IDS)
    with patch(
        "anchor.api.read_acquisition_inputs_from_bytes_with_report",
        wraps=lambda data: stub_report,
    ) as mock_read:
        response = client.post("/ingestion/excel", files=_upload_files(VALID_WORKBOOK_BYTES))

    assert response.status_code == 200
    mock_read.assert_called_once_with(VALID_WORKBOOK_BYTES)


# =============================================================================
# Upload validation -- filename / signature (KTD9)
# =============================================================================


def test_non_xlsx_filename_is_rejected_without_parsing(client: TestClient) -> None:
    with patch("anchor.api.read_acquisition_inputs_from_bytes_with_report") as mock_read:
        response = client.post(
            "/ingestion/excel",
            files={"file": ("anchor_input.csv", b"just some text", "text/csv")},
        )

    assert 400 <= response.status_code < 500
    mock_read.assert_not_called()


def test_xlsx_filename_with_non_xlsx_bytes_is_rejected_without_parsing(client: TestClient) -> None:
    with patch("anchor.api.read_acquisition_inputs_from_bytes_with_report") as mock_read:
        response = client.post(
            "/ingestion/excel",
            files=_upload_files(b"this is not a real xlsx file at all"),
        )

    assert 400 <= response.status_code < 500
    mock_read.assert_not_called()


# =============================================================================
# Upload validation -- size ceiling (KTD9)
# =============================================================================


def test_body_exceeding_the_size_ceiling_is_rejected_without_parsing(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr("anchor.api._MAX_EXCEL_UPLOAD_BYTES", 10)

    with patch("anchor.api.read_acquisition_inputs_from_bytes_with_report") as mock_read:
        response = client.post("/ingestion/excel", files=_upload_files(VALID_WORKBOOK_BYTES))

    assert 400 <= response.status_code < 500
    mock_read.assert_not_called()


def test_size_guard_enforces_the_excel_ceiling_independently_of_the_om_ceiling() -> None:
    """The shared ``_IngestionUploadSizeGuard`` instance carries a
    per-path limit mapping -- confirm the Excel path's own (smaller) ceiling
    is enforced by declared Content-Length without affecting the OM path."""

    import asyncio

    from anchor.api import _IngestionUploadSizeGuard

    async def downstream_app(scope: dict[str, Any], receive: Any, send: Any) -> None:
        await send({"type": "http.response.start", "status": 200, "headers": []})
        await send({"type": "http.response.body", "body": b"ok"})

    guard = _IngestionUploadSizeGuard(
        downstream_app, limits={"/ingestion/om": 1_000, "/ingestion/excel": 10}
    )

    async def receive() -> dict[str, Any]:
        return {"type": "http.request", "body": b"", "more_body": False}

    for path, declared_length, expected_status in (
        ("/ingestion/excel", 999, 413),
        ("/ingestion/om", 999, 200),
    ):
        sent_messages: list[dict[str, Any]] = []

        async def send(message: dict[str, Any]) -> None:
            sent_messages.append(message)

        scope = {
            "type": "http",
            "method": "POST",
            "path": path,
            "headers": [(b"content-length", str(declared_length).encode())],
        }
        asyncio.run(guard(scope, receive, send))

        start_message = next(m for m in sent_messages if m["type"] == "http.response.start")
        assert start_message["status"] == expected_status, path


# =============================================================================
# Malformed workbook -> 422 with the same issue shape /analyze uses
# =============================================================================


def test_malformed_workbook_returns_422_with_issue_list_and_never_reaches_the_engine(
    client: TestClient,
) -> None:
    values = dict(VALUES)
    values["purchase_price"] = None  # blank required value

    with patch("anchor.api.analyze_acquisition") as mock_analyze:
        response = client.post(
            "/ingestion/excel", files=_upload_files(_build_workbook_bytes(values=values))
        )

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert isinstance(detail, list)
    assert len(detail) == 1
    assert detail[0]["field_id"] == "purchase_price"
    assert detail[0]["category"] == IssueCategory.BLANK_VALUE.value
    mock_analyze.assert_not_called()


def test_workbook_missing_the_inputs_sheet_returns_422(client: TestClient) -> None:
    workbook = Workbook()
    workbook.active.title = "NotInputs"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    response = client.post("/ingestion/excel", files=_upload_files(buffer.getvalue()))

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert len(detail) == 1
    assert detail[0]["category"] == IssueCategory.MISSING_SHEET.value


def test_422_response_shape_matches_the_analyze_endpoint(client: TestClient) -> None:
    """The Excel and /analyze endpoints raise the exact same
    InputValidationError type through the exact same _validation_error_detail
    helper -- confirm the frontend never needs a second error shape."""

    values = dict(VALUES)
    values["occupancy"] = 1.5  # out of domain

    excel_response = client.post(
        "/ingestion/excel", files=_upload_files(_build_workbook_bytes(values=values))
    )
    analyze_payload = {**VALUES, "occupancy": 1.5}
    analyze_response = client.post("/analyze", json=analyze_payload)

    assert excel_response.status_code == analyze_response.status_code == 422
    excel_issue = excel_response.json()["detail"][0]
    analyze_issue = analyze_response.json()["detail"][0]
    assert set(excel_issue.keys()) == set(analyze_issue.keys()) == {
        "field_id",
        "category",
        "message",
    }
    assert excel_issue["field_id"] == analyze_issue["field_id"] == "occupancy"
    assert excel_issue["category"] == analyze_issue["category"] == IssueCategory.OUT_OF_DOMAIN_VALUE.value


# =============================================================================
# Excel upload never auto-runs analysis; other endpoints unaffected
# =============================================================================


def test_excel_upload_never_calls_the_deterministic_engine(client: TestClient) -> None:
    with patch("anchor.api.analyze_acquisition") as mock_analyze:
        response = client.post("/ingestion/excel", files=_upload_files(VALID_WORKBOOK_BYTES))

    assert response.status_code == 200
    mock_analyze.assert_not_called()


def test_analyze_endpoint_still_works_after_excel_ingestion_added(client: TestClient) -> None:
    response = client.post("/analyze", json=VALUES)

    assert response.status_code == 200


def test_om_ingestion_endpoint_still_registered_after_excel_ingestion_added(
    client: TestClient,
) -> None:
    with patch("anchor.api.extract_om") as mock_extract:
        response = client.post(
            "/ingestion/om",
            files={"file": ("om.pdf", b"not a real pdf", "application/pdf")},
        )

    # Rejected for not being a real PDF, but the route exists and is reachable.
    assert 400 <= response.status_code < 500
    mock_extract.assert_not_called()


# =============================================================================
# Underwriting V2 Gate 5 -- Excel values sent to /analyze reproduce direct
# engine analysis (both the legacy-nine and the complete-fourteen case).
# =============================================================================


def _v2_workbook_bytes() -> bytes:
    v2_values: dict[str, int | float] = {
        "acquisition_cost_pct": 0.02,
        "financing_fee_pct": 0.01,
        "disposition_cost_pct": 0.025,
        "annual_capex_reserve": 50_000,
        "io_period": 2,
    }
    v2_labels = {
        "acquisition_cost_pct": "Acquisition Cost %",
        "financing_fee_pct": "Financing Fee %",
        "disposition_cost_pct": "Disposition Cost %",
        "annual_capex_reserve": "Annual CapEx Reserve",
        "io_period": "Interest-Only Period",
    }
    v2_units = {
        "acquisition_cost_pct": "%",
        "financing_fee_pct": "%",
        "disposition_cost_pct": "%",
        "annual_capex_reserve": "USD/year",
        "io_period": "years",
    }
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet.append(HEADERS)
    for field_id in FIELD_ORDER:
        worksheet.append((field_id, LABELS[field_id], VALUES[field_id], UNITS[field_id]))
    for field_id, value in v2_values.items():
        worksheet.append((field_id, v2_labels[field_id], value, v2_units[field_id]))
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_legacy_workbook_excel_values_sent_to_analyze_reproduce_direct_engine_analysis(
    client: TestClient,
) -> None:
    from anchor.engine import analyze_acquisition

    excel_response = client.post("/ingestion/excel", files=_upload_files(VALID_WORKBOOK_BYTES))
    assert excel_response.status_code == 200
    excel_inputs = excel_response.json()["inputs"]

    analyze_response = client.post("/analyze", json=excel_inputs)
    assert analyze_response.status_code == 200

    direct_result = analyze_acquisition(EXPECTED)
    assert analyze_response.json()["levered_irr"] == pytest.approx(direct_result.levered_irr)
    assert analyze_response.json()["equity_multiple"] == pytest.approx(
        direct_result.equity_multiple
    )
    assert analyze_response.json()["loan_amount"] == direct_result.loan_amount


def test_complete_v2_workbook_excel_values_sent_to_analyze_reproduce_direct_engine_analysis(
    client: TestClient,
) -> None:
    from anchor.contracts import AcquisitionInputs as _AcquisitionInputs
    from anchor.engine import analyze_acquisition

    excel_response = client.post(
        "/ingestion/excel", files=_upload_files(_v2_workbook_bytes())
    )
    assert excel_response.status_code == 200
    body = excel_response.json()
    assert body["defaulted_v2_field_ids"] == []
    excel_inputs = body["inputs"]

    analyze_response = client.post("/analyze", json=excel_inputs)
    assert analyze_response.status_code == 200

    direct_inputs = _AcquisitionInputs(
        **{k: v for k, v in VALUES.items()},
        acquisition_cost_pct=0.02,
        financing_fee_pct=0.01,
        disposition_cost_pct=0.025,
        annual_capex_reserve=50_000.0,
        io_period=2,
    )
    direct_result = analyze_acquisition(direct_inputs)
    assert analyze_response.json()["levered_irr"] == pytest.approx(direct_result.levered_irr)
    assert analyze_response.json()["unlevered_irr"] == pytest.approx(
        direct_result.unlevered_irr
    )
    assert analyze_response.json()["min_dscr"] == pytest.approx(direct_result.min_dscr)
