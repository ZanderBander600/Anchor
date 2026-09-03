"""Detailed Operating Model V2.1 Gate 10 -- Excel ingestion for Detailed
Underwrite (``anchor.detailed_excel_reader``, ``anchor.workbook_schema``).

Mirrors ``test_excel_reader.py``'s style for the Detailed workbook's own
field set (``AcquisitionTerms`` + ``DetailedOperatingInputs``, 22 Field IDs,
all required), plus the Gate 10 workbook-schema/version metadata contract
shared with the Quick reader (``anchor.excel_reader``).
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from anchor.contracts import AcquisitionTerms, DetailedOperatingInputs
from anchor.detailed_excel_reader import (
    DetailedExcelIntakeReport,
    read_detailed_excel_intake,
    read_detailed_excel_intake_from_bytes,
)
from anchor.engine import analyze_detailed_acquisition_with_projection
from anchor.excel_reader import read_acquisition_inputs
from anchor.validation import (
    DETAILED_FIELD_IDS,
    TERMS_FIELD_IDS,
    InputValidationError,
    IssueCategory,
)
from anchor.workbook_schema import read_workbook_schema

HEADERS = ("Field ID", "Input", "Value", "Unit")
META_HEADERS = ("Key", "Value")

TERMS_VALUES: dict[str, int | float] = {
    "purchase_price": 10_000_000,
    "hold_period": 5,
    "exit_cap_rate": 0.065,
    "ltv": 0.60,
    "interest_rate": 0.05,
    "amortization": 30,
    "acquisition_cost_pct": 0.02,
    "financing_fee_pct": 0.01,
    "disposition_cost_pct": 0.025,
    "annual_capex_reserve": 50_000,
    "io_period": 2,
}
DETAILED_VALUES: dict[str, int | float] = {
    "gross_potential_rent": 800_000,
    "other_income": 20_000,
    "vacancy_credit_loss_pct": 0.05,
    "property_taxes": 60_000,
    "insurance": 20_000,
    "utilities": 25_000,
    "repairs_maintenance": 20_000,
    "other_operating_expenses": 16_000,
    "management_fee_pct": 0.05,
    "revenue_growth": 0.03,
    "expense_growth": 0.03,
}
ALL_VALUES: dict[str, int | float] = {**TERMS_VALUES, **DETAILED_VALUES}
FIELD_ORDER = TERMS_FIELD_IDS + DETAILED_FIELD_IDS
LABELS = {field_id: field_id.replace("_", " ").title() for field_id in FIELD_ORDER}
UNITS = {field_id: "" for field_id in FIELD_ORDER}

EXPECTED_TERMS = AcquisitionTerms(
    purchase_price=10_000_000.0,
    hold_period=5,
    exit_cap_rate=0.065,
    ltv=0.60,
    interest_rate=0.05,
    amortization=30,
    acquisition_cost_pct=0.02,
    financing_fee_pct=0.01,
    disposition_cost_pct=0.025,
    annual_capex_reserve=50_000.0,
    io_period=2,
)
EXPECTED_DETAILED_INPUTS = DetailedOperatingInputs(
    gross_potential_rent=800_000.0,
    other_income=20_000.0,
    vacancy_credit_loss_pct=0.05,
    property_taxes=60_000.0,
    insurance=20_000.0,
    utilities=25_000.0,
    repairs_maintenance=20_000.0,
    other_operating_expenses=16_000.0,
    management_fee_pct=0.05,
    revenue_growth=0.03,
    expense_growth=0.03,
)

EXAMPLE_WORKBOOK = (
    Path(__file__).resolve().parents[1] / "examples" / "anchor_detailed_input_v2_1.xlsx"
)
QUICK_WORKBOOK = Path(__file__).resolve().parents[1] / "examples" / "anchor_input_v2.xlsx"
LEGACY_QUICK_WORKBOOK = Path(__file__).resolve().parents[1] / "examples" / "anchor_input.xlsx"


def _build_workbook(
    *,
    anchor_schema: object = "detailed_acquisition",
    schema_version: object = "2.1",
    values: dict[str, object] | None = None,
    field_ids: tuple[str, ...] | None = None,
    include_meta: bool = True,
) -> Workbook:
    supplied = ALL_VALUES if values is None else values
    ids = FIELD_ORDER if field_ids is None else field_ids

    workbook = Workbook()
    meta = workbook.active
    meta.title = "Meta"
    if include_meta:
        meta.append(META_HEADERS)
        if anchor_schema is not None:
            meta.append(("anchor_schema", anchor_schema))
        if schema_version is not None:
            meta.append(("schema_version", schema_version))

    inputs = workbook.create_sheet("Inputs")
    inputs.append(HEADERS)
    for field_id in ids:
        inputs.append((field_id, LABELS[field_id], supplied[field_id], UNITS[field_id]))
    return workbook


def _bytes_of(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _canonical_bytes(**kwargs: Any) -> bytes:
    return _bytes_of(_build_workbook(**kwargs))


VALID_WORKBOOK_BYTES = _canonical_bytes()


# =============================================================================
# 1-6: canonical workbook parses; schema/version; all 22 fields; golden-case
# engine reconciliation.
# =============================================================================


def test_canonical_workbook_parses_successfully() -> None:
    report = read_detailed_excel_intake_from_bytes(VALID_WORKBOOK_BYTES)
    assert isinstance(report, DetailedExcelIntakeReport)


def test_schema_and_version_are_reported() -> None:
    report = read_detailed_excel_intake_from_bytes(VALID_WORKBOOK_BYTES)
    assert report.anchor_schema == "detailed_acquisition"
    assert report.schema_version == "2.1"


def test_all_eleven_acquisition_terms_values_parse_correctly() -> None:
    report = read_detailed_excel_intake_from_bytes(VALID_WORKBOOK_BYTES)
    assert report.terms == EXPECTED_TERMS


def test_all_eleven_detailed_operating_inputs_values_parse_correctly() -> None:
    report = read_detailed_excel_intake_from_bytes(VALID_WORKBOOK_BYTES)
    assert report.detailed_operating_inputs == EXPECTED_DETAILED_INPUTS


def test_tracked_example_workbook_matches_expected_contracts() -> None:
    report = read_detailed_excel_intake(EXAMPLE_WORKBOOK)
    assert report.terms == EXPECTED_TERMS
    assert report.detailed_operating_inputs == EXPECTED_DETAILED_INPUTS
    assert report.anchor_schema == "detailed_acquisition"
    assert report.schema_version == "2.1"


def test_parsed_assumptions_reproduce_the_detailed_golden_case_through_the_real_engine() -> None:
    """Never hardcodes result values into ingestion logic -- feeds the
    parser's real output into the real, unmodified engine and checks the
    engine's own output against the frozen Detailed V2.1 bridge golden
    case."""

    report = read_detailed_excel_intake(EXAMPLE_WORKBOOK)
    outcome = analyze_detailed_acquisition_with_projection(
        report.terms, report.detailed_operating_inputs
    )

    noi_by_year = outcome.operating_projection.noi_by_year
    assert noi_by_year[0] == pytest.approx(600_000.0)
    assert list(noi_by_year) == pytest.approx(
        [600_000.0, 618_000.0, 636_540.0, 655_636.2, 675_305.286]
    )
    assert outcome.operating_projection.exit_noi == pytest.approx(695_564.44458)

    results = outcome.results
    assert results.levered_irr == pytest.approx(0.073802, abs=1e-6)
    assert results.unlevered_irr == pytest.approx(0.061388, abs=1e-6)
    assert results.equity_multiple == pytest.approx(1.38235, abs=1e-5)
    assert results.dscr_by_year[0] == pytest.approx(2.00, abs=1e-6)
    assert results.min_dscr == pytest.approx(1.64688, abs=1e-5)


# =============================================================================
# 7-8: unsupported schema / unsupported version.
# =============================================================================


def test_unsupported_anchor_schema_is_rejected() -> None:
    workbook_bytes = _canonical_bytes(anchor_schema="some_other_schema")
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake_from_bytes(workbook_bytes)
    (issue,) = excinfo.value.issues
    assert issue.category == IssueCategory.SCHEMA_MISMATCH
    assert "some_other_schema" in issue.message


def test_unsupported_schema_version_is_rejected() -> None:
    workbook_bytes = _canonical_bytes(schema_version="1.0")
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake_from_bytes(workbook_bytes)
    (issue,) = excinfo.value.issues
    assert issue.category == IssueCategory.UNSUPPORTED_SCHEMA_VERSION
    assert "1.0" in issue.message


# =============================================================================
# 9-13: missing/blank/percentage/integer/duplicate field rejection.
# =============================================================================


def test_missing_detailed_field_is_rejected() -> None:
    field_ids = tuple(f for f in FIELD_ORDER if f != "gross_potential_rent")
    workbook_bytes = _canonical_bytes(field_ids=field_ids)
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake_from_bytes(workbook_bytes)
    categories = {issue.category for issue in excinfo.value.issues}
    field_ids_in_issues = {issue.field_id for issue in excinfo.value.issues}
    assert IssueCategory.MISSING_FIELD_ID in categories
    assert "gross_potential_rent" in field_ids_in_issues


def test_blank_detailed_field_is_rejected() -> None:
    values = dict(ALL_VALUES)
    values["property_taxes"] = None
    workbook_bytes = _canonical_bytes(values=values)
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake_from_bytes(workbook_bytes)
    (issue,) = [i for i in excinfo.value.issues if i.field_id == "property_taxes"]
    assert issue.category == IssueCategory.BLANK_VALUE


def test_invalid_percentage_is_rejected() -> None:
    values = dict(ALL_VALUES)
    values["vacancy_credit_loss_pct"] = 1.5  # out of [0, 1] domain
    workbook_bytes = _canonical_bytes(values=values)
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake_from_bytes(workbook_bytes)
    (issue,) = [i for i in excinfo.value.issues if i.field_id == "vacancy_credit_loss_pct"]
    assert issue.category == IssueCategory.OUT_OF_DOMAIN_VALUE


def test_invalid_integer_io_period_is_rejected() -> None:
    values = dict(ALL_VALUES)
    values["io_period"] = 2.5
    workbook_bytes = _canonical_bytes(values=values)
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake_from_bytes(workbook_bytes)
    (issue,) = [i for i in excinfo.value.issues if i.field_id == "io_period"]
    assert issue.category == IssueCategory.NON_WHOLE_NUMBER_IO_PERIOD


def test_duplicate_field_id_is_rejected() -> None:
    workbook = _build_workbook()
    inputs = workbook["Inputs"]
    inputs.append(("purchase_price", "Purchase Price (dup)", 1, ""))
    workbook_bytes = _bytes_of(workbook)
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake_from_bytes(workbook_bytes)
    (issue,) = [i for i in excinfo.value.issues if i.category == IssueCategory.DUPLICATE_FIELD_ID]
    assert issue.field_id == "purchase_price"


# =============================================================================
# 14-15: wrong-workbook protection, both directions.
# =============================================================================


def test_quick_workbook_through_detailed_path_is_rejected() -> None:
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake(QUICK_WORKBOOK)
    (issue,) = excinfo.value.issues
    assert issue.category == IssueCategory.SCHEMA_MISMATCH
    assert "Quick Underwrite" in issue.message


def test_legacy_schema_less_quick_workbook_through_detailed_path_is_rejected() -> None:
    with pytest.raises(InputValidationError) as excinfo:
        read_detailed_excel_intake(LEGACY_QUICK_WORKBOOK)
    (issue,) = excinfo.value.issues
    assert issue.category == IssueCategory.SCHEMA_MISMATCH


def test_detailed_workbook_through_quick_path_is_rejected() -> None:
    with pytest.raises(InputValidationError) as excinfo:
        read_acquisition_inputs(EXAMPLE_WORKBOOK)
    (issue,) = excinfo.value.issues
    assert issue.category == IssueCategory.SCHEMA_MISMATCH
    assert "Detailed Underwrite" in issue.message


def test_workbook_classification_cannot_silently_reinterpret_detailed_as_quick() -> None:
    """A Detailed workbook is never partially/coincidentally accepted by the
    Quick reader -- it fails before any Quick field-level parsing runs."""

    with pytest.raises(InputValidationError) as excinfo:
        read_acquisition_inputs(EXAMPLE_WORKBOOK)
    assert len(excinfo.value.issues) == 1
    assert excinfo.value.issues[0].category == IssueCategory.SCHEMA_MISMATCH


# =============================================================================
# 16-18: legacy/V2 Quick workbooks remain unaffected by Gate 10.
# =============================================================================


def test_legacy_nine_field_quick_workbook_still_works() -> None:
    inputs = read_acquisition_inputs(LEGACY_QUICK_WORKBOOK)
    assert inputs.purchase_price == 50_000_000.0


def test_quick_v2_fourteen_field_workbook_still_works() -> None:
    inputs = read_acquisition_inputs(QUICK_WORKBOOK)
    assert inputs.purchase_price is not None


def test_read_workbook_schema_returns_none_for_a_workbook_with_no_meta_sheet() -> None:
    workbook = load_workbook(LEGACY_QUICK_WORKBOOK)
    try:
        schema = read_workbook_schema(workbook)
    finally:
        workbook.close()
    assert schema.anchor_schema is None
    assert schema.schema_version is None


# =============================================================================
# 19-20: parser never calculates NOI or acquisition results.
# =============================================================================


def test_detailed_excel_intake_report_carries_no_calculated_fields() -> None:
    report = read_detailed_excel_intake(EXAMPLE_WORKBOOK)
    report_field_names = set(DetailedExcelIntakeReport.__dataclass_fields__)
    assert report_field_names == {
        "terms",
        "detailed_operating_inputs",
        "anchor_schema",
        "schema_version",
    }
    # AcquisitionTerms/DetailedOperatingInputs themselves carry no derived
    # field either (noi_by_year, exit_noi, levered_irr, etc. all absent).
    assert not hasattr(report.terms, "noi_by_year")
    assert not hasattr(report.detailed_operating_inputs, "noi_by_year")
    assert not hasattr(report, "operating_projection")
    assert not hasattr(report, "results")


# =============================================================================
# 21: deterministic error ordering (malformed/unknown rows, duplicates,
# missing, then value issues -- same order the Quick reader guarantees).
# =============================================================================


def test_error_ordering_is_deterministic_across_repeated_runs() -> None:
    values = dict(ALL_VALUES)
    values["property_taxes"] = None
    values["vacancy_credit_loss_pct"] = 1.5
    field_ids = tuple(f for f in FIELD_ORDER if f != "gross_potential_rent")
    workbook_bytes = _canonical_bytes(values=values, field_ids=field_ids)

    def _issue_signature() -> tuple[tuple[str, str | None], ...]:
        with pytest.raises(InputValidationError) as excinfo:
            read_detailed_excel_intake_from_bytes(workbook_bytes)
        return tuple((issue.category.value, issue.field_id) for issue in excinfo.value.issues)

    first = _issue_signature()
    second = _issue_signature()
    assert first == second
    categories = [category for category, _ in first]
    assert categories.index(IssueCategory.MISSING_FIELD_ID.value) < categories.index(
        IssueCategory.OUT_OF_DOMAIN_VALUE.value
    )
