from __future__ import annotations

from collections.abc import Callable, Iterable
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook

import anchor.excel_reader as excel_reader
from anchor.contracts import AcquisitionInputs
from anchor.excel_reader import read_acquisition_inputs, read_acquisition_inputs_from_bytes
from anchor.validation import (
    FIELD_IDS,
    InputIssue,
    InputValidationError,
    IssueCategory,
)


HEADERS = ("Field ID", "Input", "Value", "Unit")
VALUES: dict[str, int | float] = {
    "purchase_price": 50_000_000,
    "current_noi": 2_500_000,
    "occupancy": 0.95,
    "noi_growth": 0.03,
    "hold_period": 5,
    "exit_cap_rate": 0.055,
    "ltv": 0.65,
    "interest_rate": 0.0525,
    "amortization": 30,
}
LABELS = {
    "purchase_price": "Purchase Price",
    "current_noi": "Current NOI",
    "occupancy": "Occupancy",
    "noi_growth": "NOI Growth",
    "hold_period": "Hold Period",
    "exit_cap_rate": "Exit Cap Rate",
    "ltv": "LTV",
    "interest_rate": "Interest Rate",
    "amortization": "Amortization",
}
UNITS = {
    "purchase_price": "USD",
    "current_noi": "USD/year",
    "occupancy": "%",
    "noi_growth": "%/year",
    "hold_period": "years",
    "exit_cap_rate": "%",
    "ltv": "%",
    "interest_rate": "%",
    "amortization": "years",
}
EXPECTED = AcquisitionInputs(
    purchase_price=50_000_000.0,
    current_noi=2_500_000.0,
    occupancy=0.95,
    noi_growth=0.03,
    hold_period=5,
    exit_cap_rate=0.055,
    ltv=0.65,
    interest_rate=0.0525,
    amortization=30,
)
EXAMPLE_WORKBOOK = Path(__file__).resolve().parents[1] / "examples" / "anchor_input.xlsx"


def _canonical_rows(
    values: dict[str, object] | None = None,
) -> list[tuple[object, object, object, object]]:
    supplied = VALUES if values is None else values
    return [
        (field_id, LABELS[field_id], supplied[field_id], UNITS[field_id])
        for field_id in FIELD_IDS
    ]


def _populate_workbook(
    *,
    rows: Iterable[tuple[object, object, object, object]] | None = None,
    headers: Iterable[object] = HEADERS,
    sheet_title: str = "Inputs",
) -> tuple[Workbook, Any]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    worksheet.append(tuple(headers))
    for row in _canonical_rows() if rows is None else rows:
        worksheet.append(tuple(row))
    return workbook, worksheet


def _write_workbook(
    tmp_path: Path,
    *,
    rows: Iterable[tuple[object, object, object, object]] | None = None,
    headers: Iterable[object] = HEADERS,
    sheet_title: str = "Inputs",
    configure: Callable[[Workbook, Any], None] | None = None,
    filename: str = "input.xlsx",
) -> Path:
    workbook, worksheet = _populate_workbook(
        rows=rows, headers=headers, sheet_title=sheet_title
    )
    if configure is not None:
        configure(workbook, worksheet)
    path = tmp_path / filename
    workbook.save(path)
    workbook.close()
    return path


def _workbook_with_value(
    tmp_path: Path,
    field_id: str,
    value: object,
    *,
    configure: Callable[[Workbook, Any], None] | None = None,
) -> Path:
    values: dict[str, object] = dict(VALUES)
    values[field_id] = value
    return _write_workbook(tmp_path, rows=_canonical_rows(values), configure=configure)


def _capture_issues(path: Path) -> tuple[InputIssue, ...]:
    with pytest.raises(InputValidationError) as captured:
        read_acquisition_inputs(path)
    return captured.value.issues


def _single_issue(path: Path) -> InputIssue:
    issues = _capture_issues(path)
    assert len(issues) == 1
    return issues[0]


def _cache_formula_result(path: Path, coordinate: str, value: float) -> None:
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ElementTree.register_namespace("", namespace)
    rewritten = path.with_name("cached-formula.xlsx")

    with ZipFile(path, "r") as source, ZipFile(rewritten, "w") as destination:
        for member in source.infolist():
            content = source.read(member.filename)
            if member.filename == "xl/worksheets/sheet1.xml":
                root = ElementTree.fromstring(content)
                for cell in root.findall(f".//{{{namespace}}}c"):
                    if cell.get("r") != coordinate:
                        continue
                    assert cell.find(f"{{{namespace}}}f") is not None
                    cached = cell.find(f"{{{namespace}}}v")
                    if cached is None:
                        cached = ElementTree.SubElement(cell, f"{{{namespace}}}v")
                    cached.text = str(value)
                    break
                else:
                    raise AssertionError(f"Formula cell {coordinate} was not found")
                content = ElementTree.tostring(root, encoding="utf-8")
            destination.writestr(member, content)

    rewritten.replace(path)


def test_valid_canonical_workbook_returns_exact_inputs_and_types(tmp_path: Path) -> None:
    result = read_acquisition_inputs(_write_workbook(tmp_path))

    assert result == EXPECTED
    for field_id in FIELD_IDS:
        expected_type = int if field_id in {"hold_period", "amortization"} else float
        assert type(getattr(result, field_id)) is expected_type


def test_tracked_example_workbook_has_exact_raw_values_and_percentage_formats() -> None:
    workbook = load_workbook(EXAMPLE_WORKBOOK, data_only=False)
    try:
        assert workbook.sheetnames == ["Inputs"]
        worksheet = workbook["Inputs"]
        assert tuple(worksheet.cell(row=1, column=column).value for column in range(1, 5)) == HEADERS
        raw = {
            worksheet.cell(row=row_number, column=1).value: worksheet.cell(
                row=row_number, column=3
            ).value
            for row_number in range(2, 11)
        }
        formats = {
            worksheet.cell(row=row_number, column=1).value: worksheet.cell(
                row=row_number, column=3
            ).number_format
            for row_number in range(2, 11)
        }
    finally:
        workbook.close()

    assert raw == VALUES
    for field_id in (
        "occupancy",
        "noi_growth",
        "exit_cap_rate",
        "ltv",
        "interest_rate",
    ):
        assert formats[field_id] == "0.00%"
    assert read_acquisition_inputs(EXAMPLE_WORKBOOK) == EXPECTED


def test_row_order_does_not_change_result(tmp_path: Path) -> None:
    rows = list(reversed(_canonical_rows()))
    assert read_acquisition_inputs(_write_workbook(tmp_path, rows=rows)) == EXPECTED


def test_input_labels_and_unit_text_are_descriptive_only(tmp_path: Path) -> None:
    rows = [
        (field_id, f"arbitrary label {index}", VALUES[field_id], f"unit {index}")
        for index, field_id in enumerate(FIELD_IDS)
    ]
    rows[0] = ("purchase_price", "=1+1", VALUES["purchase_price"], "=2+2")

    assert read_acquisition_inputs(_write_workbook(tmp_path, rows=rows)) == EXPECTED


def test_number_formats_styles_hidden_rows_and_filters_do_not_change_values(
    tmp_path: Path,
) -> None:
    def configure(_workbook: Workbook, worksheet: Any) -> None:
        worksheet["C2"].number_format = '$#,##0.00;[Red]-$#,##0.00'
        worksheet["C3"].number_format = '$#,##0.00'
        for row in (4, 5, 7, 8, 9):
            worksheet.cell(row=row, column=3).number_format = "0.00%"
        worksheet.row_dimensions[4].hidden = True
        worksheet.auto_filter.ref = "A1:D10"

    path = _write_workbook(tmp_path, configure=configure)
    result = read_acquisition_inputs(path)

    assert result == EXPECTED
    assert result.occupancy == 0.95
    assert result.noi_growth == 0.03
    assert result.exit_cap_rate == 0.055
    assert result.ltv == 0.65
    assert result.interest_rate == 0.0525


def test_occupancy_is_preserved_and_never_applied_to_current_noi(tmp_path: Path) -> None:
    values: dict[str, object] = dict(VALUES)
    values["occupancy"] = 0.4
    values["current_noi"] = 2_345_678

    result = read_acquisition_inputs(
        _write_workbook(tmp_path, rows=_canonical_rows(values))
    )

    assert result.occupancy == 0.4
    assert result.current_noi == 2_345_678.0


def test_underlying_occupancy_95_is_not_rescaled_and_is_rejected(tmp_path: Path) -> None:
    issue = _single_issue(_workbook_with_value(tmp_path, "occupancy", 95))

    assert issue.category is IssueCategory.OUT_OF_DOMAIN_VALUE
    assert issue.field_id == "occupancy"
    assert issue.value == 95.0


def test_literal_integer_year_values_produce_python_ints(tmp_path: Path) -> None:
    result = read_acquisition_inputs(_write_workbook(tmp_path))

    assert type(result.hold_period) is int
    assert result.hold_period == 5
    assert type(result.amortization) is int
    assert result.amortization == 30


@pytest.mark.parametrize(
    ("field_id", "value"), (("hold_period", 5.0), ("amortization", 30.0))
)
def test_integral_float_year_values_normalize_to_int(
    monkeypatch: pytest.MonkeyPatch, field_id: str, value: float
) -> None:
    values: dict[str, object] = dict(VALUES)
    values[field_id] = value
    workbook, worksheet = _populate_workbook(rows=_canonical_rows(values))
    value_cell = worksheet.cell(row=FIELD_IDS.index(field_id) + 2, column=3)
    assert type(value_cell.value) is float

    def fake_load_workbook(_path: object, *, data_only: bool) -> Workbook:
        assert data_only is False
        return workbook

    monkeypatch.setattr(excel_reader, "load_workbook", fake_load_workbook)
    result = read_acquisition_inputs("in-memory.xlsx")

    assert type(getattr(result, field_id)) is int
    assert getattr(result, field_id) == int(value)


@pytest.mark.parametrize(
    ("field_id", "value", "category"),
    (
        ("hold_period", 5.5, IssueCategory.NON_WHOLE_NUMBER_HOLD_PERIOD),
        ("amortization", 30.25, IssueCategory.NON_WHOLE_NUMBER_AMORTIZATION),
    ),
)
def test_fractional_year_values_receive_field_specific_errors(
    tmp_path: Path, field_id: str, value: float, category: IssueCategory
) -> None:
    issue = _single_issue(_workbook_with_value(tmp_path, field_id, value))

    assert issue.category is category
    assert issue.field_id == field_id
    assert issue.row == FIELD_IDS.index(field_id) + 2
    assert issue.cell == f"C{issue.row}"


@pytest.mark.parametrize("sheet_title", ("inputs", "INPUTS", "Inputs "))
def test_inputs_sheet_name_is_exact_and_case_sensitive(
    tmp_path: Path, sheet_title: str
) -> None:
    issue = _single_issue(_write_workbook(tmp_path, sheet_title=sheet_title))

    assert issue.category is IssueCategory.MISSING_SHEET
    assert issue.field_id is None
    assert "Inputs" in issue.message


def test_other_worksheets_are_ignored(tmp_path: Path) -> None:
    def configure(workbook: Workbook, _worksheet: Any) -> None:
        other = workbook.create_sheet("Other")
        other.append(HEADERS)
        other.append(("unknown", "Bad", "not numeric", "none"))

    assert read_acquisition_inputs(_write_workbook(tmp_path, configure=configure)) == EXPECTED


@pytest.mark.parametrize(
    "headers",
    (
        (None, "Input", "Value", "Unit"),
        ("Field Identifier", "Input", "Value", "Unit"),
        ("field ID", "Input", "Value", "Unit"),
        ("Field ID ", "Input", "Value", "Unit"),
        ("Field ID", "Value", "Input", "Unit"),
        ("Field ID", "Input", "Input", "Unit"),
        (None, "Field ID", "Input", "Value", "Unit"),
    ),
)
def test_missing_renamed_case_changed_whitespace_shifted_duplicated_or_reordered_headers_fail(
    tmp_path: Path, headers: tuple[object, ...]
) -> None:
    issue = _single_issue(_write_workbook(tmp_path, headers=headers))

    assert issue.category is IssueCategory.MALFORMED_TABLE
    assert issue.row == 1
    assert issue.cell in {"A1", "B1", "C1", "D1"}


@pytest.mark.parametrize("merged_range", ("A1:B1", "A1:A2", "D1:E1"))
def test_any_merge_intersecting_required_header_range_is_malformed(
    tmp_path: Path, merged_range: str
) -> None:
    def configure(_workbook: Workbook, worksheet: Any) -> None:
        worksheet.merge_cells(merged_range)

    issue = _single_issue(_write_workbook(tmp_path, configure=configure))

    assert issue.category is IssueCategory.MALFORMED_TABLE
    assert issue.cell == merged_range


def test_multiple_header_merges_choose_deterministic_first_range(tmp_path: Path) -> None:
    def configure(_workbook: Workbook, worksheet: Any) -> None:
        worksheet.merge_cells("C1:D1")
        worksheet.merge_cells("A1:A2")

    issue = _single_issue(_write_workbook(tmp_path, configure=configure))

    assert issue.category is IssueCategory.MALFORMED_TABLE
    assert issue.cell == "A1:A2"


@pytest.mark.parametrize("missing_field_id", FIELD_IDS)
def test_each_missing_required_field_id_is_identified(
    tmp_path: Path, missing_field_id: str
) -> None:
    rows = [row for row in _canonical_rows() if row[0] != missing_field_id]
    issue = _single_issue(_write_workbook(tmp_path, rows=rows))

    assert issue.category is IssueCategory.MISSING_FIELD_ID
    assert issue.field_id == missing_field_id


@pytest.mark.parametrize("duplicate_field_id", FIELD_IDS)
def test_each_duplicate_required_field_id_identifies_all_rows(
    tmp_path: Path, duplicate_field_id: str
) -> None:
    rows = _canonical_rows()
    rows.append(
        (
            duplicate_field_id,
            "Duplicate",
            VALUES[duplicate_field_id],
            "ignored unit",
        )
    )
    issue = _single_issue(_write_workbook(tmp_path, rows=rows))

    assert issue.category is IssueCategory.DUPLICATE_FIELD_ID
    assert issue.field_id == duplicate_field_id
    assert issue.rows == (FIELD_IDS.index(duplicate_field_id) + 2, 11)


def test_three_duplicate_occurrences_identify_every_conflicting_row(
    tmp_path: Path,
) -> None:
    rows = _canonical_rows()
    rows.extend(
        (
            ("purchase_price", "Duplicate two", 50_000_000, "USD"),
            ("purchase_price", "Duplicate three", 50_000_000, "USD"),
        )
    )
    issue = _single_issue(_write_workbook(tmp_path, rows=rows))

    assert issue.category is IssueCategory.DUPLICATE_FIELD_ID
    assert issue.field_id == "purchase_price"
    assert issue.rows == (2, 11, 12)


def test_unknown_additional_field_id_is_identified(tmp_path: Path) -> None:
    rows = _canonical_rows()
    rows.append(("extra_field", "Extra", 1, "none"))
    issue = _single_issue(_write_workbook(tmp_path, rows=rows))

    assert issue.category is IssueCategory.UNKNOWN_FIELD_ID
    assert issue.field_id == "extra_field"
    assert issue.row == 11
    assert issue.cell == "A11"


@pytest.mark.parametrize(
    "supplied_id", (" purchase_price", "purchase_price ", "PURCHASE_PRICE", "purchase-price")
)
def test_nonblank_field_ids_are_not_trimmed_case_folded_or_fuzzily_matched(
    tmp_path: Path, supplied_id: str
) -> None:
    rows = _canonical_rows()
    rows[0] = (supplied_id, "Purchase Price", 50_000_000, "USD")
    issues = _capture_issues(_write_workbook(tmp_path, rows=rows))

    assert [(issue.category, issue.field_id) for issue in issues] == [
        (IssueCategory.UNKNOWN_FIELD_ID, supplied_id),
        (IssueCategory.MISSING_FIELD_ID, "purchase_price"),
    ]
    assert issues[0].row == 2


@pytest.mark.parametrize("blank_field_id", (None, "", "   ", "\t"))
def test_blank_field_id_on_nonempty_row_is_malformed_not_unknown(
    tmp_path: Path, blank_field_id: object
) -> None:
    rows = _canonical_rows()
    rows.append((blank_field_id, None, 1_000_000, None))
    issue = _single_issue(_write_workbook(tmp_path, rows=rows))

    assert issue.category is IssueCategory.MALFORMED_TABLE
    assert issue.row == 11
    assert issue.cell == "A11"


@pytest.mark.parametrize("field_id_value", (123, True, date(2026, 1, 1), "=\"purchase_price\""))
def test_formula_or_nontext_field_id_on_nonempty_row_is_malformed(
    tmp_path: Path, field_id_value: object
) -> None:
    rows = _canonical_rows()
    rows[0] = (field_id_value, "Purchase Price", 50_000_000, "USD")
    issues = _capture_issues(_write_workbook(tmp_path, rows=rows))

    assert [issue.category for issue in issues] == [
        IssueCategory.MALFORMED_TABLE,
        IssueCategory.MISSING_FIELD_ID,
    ]
    assert issues[0].row == 2
    assert issues[1].field_id == "purchase_price"


def test_empty_rows_do_not_end_scan_and_content_outside_a_through_d_is_ignored(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet.append(HEADERS)
    target_rows = (3, 4, 6, 7, 9, 10, 12, 13, 15)
    for row_number, record in zip(target_rows, _canonical_rows(), strict=True):
        for column, value in enumerate(record, start=1):
            worksheet.cell(row=row_number, column=column, value=value)
    worksheet["E1048576"] = "unknown_field_outside_table"
    worksheet["A80"].number_format = "0.00"
    path = tmp_path / "gapped.xlsx"
    workbook.save(path)
    workbook.close()

    assert read_acquisition_inputs(path) == EXPECTED


@pytest.mark.parametrize("blank_value", (None, "", "   ", "\t"))
def test_all_blank_value_representations_are_rejected(
    tmp_path: Path, blank_value: object
) -> None:
    issue = _single_issue(_workbook_with_value(tmp_path, "purchase_price", blank_value))

    assert issue.category is IssueCategory.BLANK_VALUE
    assert issue.field_id == "purchase_price"
    assert issue.row == 2
    assert issue.cell == "C2"


def test_formula_value_is_rejected(tmp_path: Path) -> None:
    issue = _single_issue(_workbook_with_value(tmp_path, "purchase_price", "=25000000*2"))

    assert issue.category is IssueCategory.FORMULA_VALUE
    assert issue.field_id == "purchase_price"
    assert issue.cell == "C2"


def test_formula_with_a_cached_numeric_result_is_still_rejected(tmp_path: Path) -> None:
    path = _workbook_with_value(tmp_path, "purchase_price", "=25000000*2")
    _cache_formula_result(path, "C2", 50_000_000)

    cached_workbook = load_workbook(path, data_only=True)
    try:
        assert cached_workbook["Inputs"]["C2"].value == 50_000_000
    finally:
        cached_workbook.close()

    issue = _single_issue(path)
    assert issue.category is IssueCategory.FORMULA_VALUE
    assert issue.field_id == "purchase_price"


@pytest.mark.parametrize(
    "value",
    (
        "1000000",
        "$1,000,000",
        "5.25%",
        True,
        False,
        date(2026, 8, 24),
        datetime(2026, 8, 24, 10, 30),
        time(10, 30),
        timedelta(hours=10, minutes=30),
    ),
)
def test_text_booleans_dates_times_and_durations_are_non_numeric(
    tmp_path: Path, value: object
) -> None:
    issue = _single_issue(_workbook_with_value(tmp_path, "purchase_price", value))

    assert issue.category is IssueCategory.NON_NUMERIC_VALUE
    assert issue.field_id == "purchase_price"


def test_excel_error_cell_is_non_numeric(tmp_path: Path) -> None:
    def configure(_workbook: Workbook, worksheet: Any) -> None:
        worksheet["C2"] = "#DIV/0!"
        worksheet["C2"].data_type = "e"

    issue = _single_issue(
        _workbook_with_value(tmp_path, "purchase_price", 1, configure=configure)
    )

    assert issue.category is IssueCategory.NON_NUMERIC_VALUE
    assert issue.field_id == "purchase_price"
    assert issue.value == "#DIV/0!"


@pytest.mark.parametrize("value", (float("nan"), float("inf"), float("-inf")))
def test_numeric_nan_and_infinities_are_non_finite(
    monkeypatch: pytest.MonkeyPatch, value: float
) -> None:
    values: dict[str, object] = dict(VALUES)
    values["purchase_price"] = value
    workbook, worksheet = _populate_workbook(rows=_canonical_rows(values))
    assert worksheet["C2"].data_type == "n"

    monkeypatch.setattr(
        excel_reader,
        "load_workbook",
        lambda _path, *, data_only: workbook,
    )
    issue = _single_issue(Path("in-memory.xlsx"))

    assert issue.category is IssueCategory.NON_FINITE_VALUE
    assert issue.field_id == "purchase_price"


def test_duplicate_records_are_not_value_validated(tmp_path: Path) -> None:
    rows = _canonical_rows()
    rows[0] = ("purchase_price", "First", "=1+1", "USD")
    rows.append(("purchase_price", "Second", "not numeric", "USD"))
    issue = _single_issue(_write_workbook(tmp_path, rows=rows))

    assert issue.category is IssueCategory.DUPLICATE_FIELD_ID
    assert issue.field_id == "purchase_price"
    assert issue.rows == (2, 11)


def test_value_issues_are_ordered_canonically_not_by_row(tmp_path: Path) -> None:
    values: dict[str, object] = dict(VALUES)
    values["purchase_price"] = 0
    values["occupancy"] = "95%"
    values["hold_period"] = 5.5
    rows = list(reversed(_canonical_rows(values)))
    issues = _capture_issues(_write_workbook(tmp_path, rows=rows))

    assert [(issue.category, issue.field_id) for issue in issues] == [
        (IssueCategory.OUT_OF_DOMAIN_VALUE, "purchase_price"),
        (IssueCategory.NON_NUMERIC_VALUE, "occupancy"),
        (IssueCategory.NON_WHOLE_NUMBER_HOLD_PERIOD, "hold_period"),
    ]


def test_multiple_error_ordering_matches_frozen_precedence(tmp_path: Path) -> None:
    rows = [
        ("unknown_z", None, 1, None),
        ("   ", None, 1, None),
        ("unknown_a", None, 1, None),
        ("current_noi", None, 2_500_000, None),
        ("purchase_price", None, "=1+1", None),
        ("current_noi", None, "not numeric", None),
        ("purchase_price", None, 50_000_000, None),
        ("occupancy", None, 2, None),
        ("hold_period", None, 5.5, None),
        ("ltv", None, "65%", None),
    ]
    path = _write_workbook(tmp_path, rows=rows)
    first = _capture_issues(path)
    second = _capture_issues(path)

    expected = [
        (IssueCategory.UNKNOWN_FIELD_ID, "unknown_z"),
        (IssueCategory.MALFORMED_TABLE, None),
        (IssueCategory.UNKNOWN_FIELD_ID, "unknown_a"),
        (IssueCategory.DUPLICATE_FIELD_ID, "purchase_price"),
        (IssueCategory.DUPLICATE_FIELD_ID, "current_noi"),
        (IssueCategory.MISSING_FIELD_ID, "noi_growth"),
        (IssueCategory.MISSING_FIELD_ID, "exit_cap_rate"),
        (IssueCategory.MISSING_FIELD_ID, "interest_rate"),
        (IssueCategory.MISSING_FIELD_ID, "amortization"),
        (IssueCategory.OUT_OF_DOMAIN_VALUE, "occupancy"),
        (IssueCategory.NON_WHOLE_NUMBER_HOLD_PERIOD, "hold_period"),
        (IssueCategory.NON_NUMERIC_VALUE, "ltv"),
    ]
    assert [(issue.category, issue.field_id) for issue in first] == expected
    assert first == second
    assert [issue.row for issue in first[:3]] == [2, 3, 4]
    assert first[3].rows == (6, 8)
    assert first[4].rows == (5, 7)
    assert all(
        issue.category is not IssueCategory.FORMULA_VALUE for issue in first
    )


def test_integral_nonpositive_year_is_domain_error_not_whole_number_error(
    tmp_path: Path,
) -> None:
    issue = _single_issue(_workbook_with_value(tmp_path, "hold_period", 0.0))

    assert issue.category is IssueCategory.OUT_OF_DOMAIN_VALUE
    assert issue.field_id == "hold_period"


def test_nonfinite_precedes_domain_validation(monkeypatch: pytest.MonkeyPatch) -> None:
    values: dict[str, object] = dict(VALUES)
    values["occupancy"] = float("inf")
    workbook, _worksheet = _populate_workbook(rows=_canonical_rows(values))
    monkeypatch.setattr(
        excel_reader,
        "load_workbook",
        lambda _path, *, data_only: workbook,
    )

    issue = _single_issue(Path("in-memory.xlsx"))
    assert issue.category is IssueCategory.NON_FINITE_VALUE
    assert issue.field_id == "occupancy"


def test_malformed_header_is_terminal_before_row_analysis(tmp_path: Path) -> None:
    rows = [("unknown", None, "=1+1", None)]
    issue = _single_issue(
        _write_workbook(
            tmp_path,
            rows=rows,
            headers=("field id", "Input", "Value", "Unit"),
        )
    )

    assert issue.category is IssueCategory.MALFORMED_TABLE
    assert issue.row == 1


def test_empty_table_reports_all_missing_ids_in_canonical_order(tmp_path: Path) -> None:
    issues = _capture_issues(_write_workbook(tmp_path, rows=[]))

    assert [issue.category for issue in issues] == [
        IssueCategory.MISSING_FIELD_ID
    ] * len(FIELD_IDS)
    assert [issue.field_id for issue in issues] == list(FIELD_IDS)


def test_missing_path_is_translated_to_workbook_open_issue(tmp_path: Path) -> None:
    path = tmp_path / "does-not-exist.xlsx"
    issue = _single_issue(path)

    assert issue.category is IssueCategory.WORKBOOK_OPEN
    assert issue.field_id is None
    assert str(path) in issue.message


def test_unrenderable_workbook_identifier_does_not_escape_error_contract() -> None:
    class HostilePath:
        def __fspath__(self) -> str:
            raise RuntimeError("cannot render path")

        def __repr__(self) -> str:
            raise RuntimeError("cannot render object")

    with pytest.raises(InputValidationError) as captured:
        read_acquisition_inputs(HostilePath())  # type: ignore[arg-type]

    issue = captured.value.issues[0]
    assert issue.category is IssueCategory.WORKBOOK_OPEN
    assert "<workbook>" in issue.message


def test_corrupt_nonworkbook_is_translated_to_workbook_open_issue(tmp_path: Path) -> None:
    path = tmp_path / "corrupt.xlsx"
    path.write_text("this is not an Excel workbook", encoding="utf-8")

    assert _single_issue(path).category is IssueCategory.WORKBOOK_OPEN


def test_unsupported_excel_extension_is_translated_to_workbook_open_issue(
    tmp_path: Path,
) -> None:
    xlsx_path = _write_workbook(tmp_path)
    unsupported_path = tmp_path / "input.xls"
    xlsx_path.rename(unsupported_path)

    assert _single_issue(unsupported_path).category is IssueCategory.WORKBOOK_OPEN


def test_openable_xlsm_is_rejected_as_noncanonical(tmp_path: Path) -> None:
    xlsx_path = _write_workbook(tmp_path)
    xlsm_path = tmp_path / "input.xlsm"
    xlsx_path.rename(xlsm_path)
    probe = load_workbook(xlsm_path, data_only=False)
    probe.close()

    assert _single_issue(xlsm_path).category is IssueCategory.WORKBOOK_OPEN


def test_openpyxl_load_exception_is_not_exposed(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_to_load(_path: object, *, data_only: bool) -> Workbook:
        assert data_only is False
        raise RuntimeError("sensitive library detail")

    monkeypatch.setattr(excel_reader, "load_workbook", fail_to_load)
    with pytest.raises(InputValidationError) as captured:
        read_acquisition_inputs("broken.xlsx")

    assert len(captured.value.issues) == 1
    assert captured.value.issues[0].category is IssueCategory.WORKBOOK_OPEN
    assert "sensitive library detail" not in str(captured.value)


def test_workbook_library_exception_after_open_is_translated(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class BrokenWorkbook:
        closed = False

        @property
        def sheetnames(self) -> list[str]:
            raise RuntimeError("sensitive parse detail")

        def close(self) -> None:
            self.closed = True

    workbook = BrokenWorkbook()
    monkeypatch.setattr(
        excel_reader,
        "load_workbook",
        lambda _path, *, data_only: workbook,
    )

    with pytest.raises(InputValidationError) as captured:
        read_acquisition_inputs("broken.xlsx")

    assert captured.value.issues[0].category is IssueCategory.WORKBOOK_OPEN
    assert "sensitive parse detail" not in str(captured.value)
    assert workbook.closed is True


# =============================================================================
# read_acquisition_inputs_from_bytes -- shares _read_acquisition_inputs_from_source
# with the path-based reader above; these tests prove behavioral parity rather
# than re-deriving the full malformed-workbook matrix already covered for the
# path-based entry point.
# =============================================================================


def _workbook_bytes(
    *,
    rows: Iterable[tuple[object, object, object, object]] | None = None,
    headers: Iterable[object] = HEADERS,
    sheet_title: str = "Inputs",
    configure: Callable[[Workbook, Any], None] | None = None,
) -> bytes:
    workbook, worksheet = _populate_workbook(rows=rows, headers=headers, sheet_title=sheet_title)
    if configure is not None:
        configure(workbook, worksheet)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_bytes_reader_returns_identical_result_to_path_reader_for_the_canonical_workbook() -> None:
    data = EXAMPLE_WORKBOOK.read_bytes()

    from_bytes = read_acquisition_inputs_from_bytes(data)
    from_path = read_acquisition_inputs(EXAMPLE_WORKBOOK)

    assert from_bytes == from_path == EXPECTED


def test_bytes_reader_returns_exact_inputs_and_types_for_a_freshly_built_workbook() -> None:
    result = read_acquisition_inputs_from_bytes(_workbook_bytes())

    assert result == EXPECTED
    for field_id in FIELD_IDS:
        expected_type = int if field_id in {"hold_period", "amortization"} else float
        assert type(getattr(result, field_id)) is expected_type


def test_bytes_reader_reports_the_same_issue_as_the_path_reader_for_a_missing_sheet(
    tmp_path: Path,
) -> None:
    data = _workbook_bytes(sheet_title="NotInputs")
    path = _write_workbook(tmp_path, sheet_title="NotInputs")

    bytes_issues = _capture_issues_from_bytes(data)
    path_issues = _capture_issues(path)

    assert len(bytes_issues) == len(path_issues) == 1
    assert bytes_issues[0].category == path_issues[0].category == IssueCategory.MISSING_SHEET
    assert bytes_issues[0].message == path_issues[0].message


def test_bytes_reader_reports_the_same_issue_as_the_path_reader_for_a_blank_value(
    tmp_path: Path,
) -> None:
    values: dict[str, object] = dict(VALUES)
    values["purchase_price"] = None
    rows = _canonical_rows(values)

    data = _workbook_bytes(rows=rows)
    path = _write_workbook(tmp_path, rows=rows)

    bytes_issues = _capture_issues_from_bytes(data)
    path_issues = _capture_issues(path)

    assert len(bytes_issues) == len(path_issues) == 1
    assert bytes_issues[0].category == path_issues[0].category == IssueCategory.BLANK_VALUE
    assert bytes_issues[0].field_id == path_issues[0].field_id == "purchase_price"
    assert bytes_issues[0].message == path_issues[0].message


def test_bytes_reader_rejects_unopenable_bytes_with_the_workbook_open_category() -> None:
    with pytest.raises(InputValidationError) as captured:
        read_acquisition_inputs_from_bytes(b"this is not a real xlsx file at all")

    issues = captured.value.issues
    assert len(issues) == 1
    assert issues[0].category is IssueCategory.WORKBOOK_OPEN
    assert "<uploaded workbook>" in issues[0].message


def test_bytes_reader_does_not_require_or_check_a_filename_extension() -> None:
    """Unlike read_acquisition_inputs, the bytes entry point has no filename
    to gate on -- valid workbook bytes are accepted regardless of any
    filename the caller might otherwise have had for the upload."""

    result = read_acquisition_inputs_from_bytes(_workbook_bytes())

    assert result == EXPECTED


def _capture_issues_from_bytes(data: bytes) -> tuple[InputIssue, ...]:
    with pytest.raises(InputValidationError) as captured:
        read_acquisition_inputs_from_bytes(data)
    return captured.value.issues
